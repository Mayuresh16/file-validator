# Licensed to the Apache Software Foundation (ASF) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The ASF licenses this file
# to you under the MIT License (the "License"); you may not
# use this file except in compliance with the License.
#
# MIT License
#
# Copyright (c) 2026 Mayuresh Kedari
#
# Permission is hereby granted, free of charge, to any person
# obtaining a copy of this software and associated documentation
# files (the "Software"), to deal in the Software without
# restriction, including without limitation the rights to use,
# copy, modify, merge, publish, distribute, sublicense, and/or
# sell copies of the Software, and to permit persons to whom
# the Software is furnished to do so, subject to the following
# conditions:
#
# The above copyright notice and this permission notice shall be
# included in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY
# KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE
# WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR
# PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS
# OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR
# OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR
# OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE
# SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.

"""Excel export functionality for file validation results."""

import logging
from datetime import datetime
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from file_validator.constants import HeaderColor, StatusColor, TextColor

logger: logging.Logger = logging.getLogger(__name__)


# ============================================================
# Helper Functions
# ============================================================


def _apply_header_style(
    cell, fill_color: str = HeaderColor.LIGHT_BLUE, text_color: str = TextColor.BLACK, bold: bool = True
) -> None:
    """Apply consistent header styling to a cell."""
    cell.font = Font(bold=bold, color=text_color)
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")


def _write_header_row(
    ws, headers: list[str], row_num: int = 1, fill_color: str = HeaderColor.LIGHT_BLUE
) -> None:
    """Write header row with consistent styling."""
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row_num, col_num, header)
        _apply_header_style(ws.cell(row_num, col_num), fill_color=fill_color)


def _write_data_rows(
    ws, df: pl.DataFrame, columns: list[str], start_row: int = 2, max_rows: int = 10000
) -> int:
    """Write data rows to worksheet and return the number of rows written."""
    row = start_row
    for record in df.iter_rows(named=True):
        for col_num, col_name in enumerate(columns, start=1):
            ws.cell(row, col_num, record.get(col_name, ""))
        row += 1
        if row > start_row + max_rows:
            break
    return row - start_row


def _auto_adjust_columns(ws, num_cols: int = 0, default_width: int = 15) -> None:
    """Auto-adjust column widths."""
    if num_cols == 0:
        num_cols = ws.max_column
    for col_num in range(1, num_cols + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = default_width


def _extract_data_columns(
    df: pl.DataFrame, sample_df: pl.DataFrame | None, primary_keys: list[str]
) -> list[str]:
    """Extract data column names from comparison DataFrame."""
    working_df = sample_df if sample_df is not None and not sample_df.is_empty() else df
    status_cols = [c for c in working_df.columns if c.endswith("_status")]
    data_cols = [c.replace("_status", "") for c in status_cols if c != "validation_status"]

    if not data_cols:
        all_cols = [c for c in working_df.columns if not c.endswith(("_source", "_target", "_status"))]
        data_cols = [c for c in all_cols if c not in primary_keys and c != "validation_status"]

    return data_cols


def _create_line_comparison_sheet(wb: Workbook, sheet_name: str, line_comparison: list[dict]) -> None:
    """Create a line comparison sheet (header or trailer) with consistent structure."""
    ws = wb.create_sheet(sheet_name)

    # Headers
    headers = ["Line", "Status", "Source", "Target"]
    _write_header_row(ws, headers, row_num=1, fill_color=HeaderColor.LIGHT_BLUE)

    # Data
    for row_num, record in enumerate(line_comparison, start=2):
        ws.cell(row_num, 1, record.get("line_number", ""))
        ws.cell(row_num, 2, record.get("status", ""))
        ws.cell(row_num, 3, record.get("source", ""))
        ws.cell(row_num, 4, record.get("target", ""))

        status: str = record.get("status", "")
        if status == "MATCH":
            ws.cell(row_num, 2).fill = PatternFill(
                start_color=StatusColor.MATCH, end_color=StatusColor.MATCH, fill_type="solid"
            )
        elif "Mismatch" in status:
            ws.cell(row_num, 2).fill = PatternFill(
                start_color=StatusColor.MISMATCH, end_color=StatusColor.MISMATCH, fill_type="solid"
            )

    _auto_adjust_columns(ws, num_cols=4, default_width=20)


def _create_comparison_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pl.DataFrame,
    primary_keys: list[str],
    data_cols: list[str],
    suffix: str,
    header_color: str,
) -> None:
    """Create a data sheet (source or target) with consistent structure."""
    ws = wb.create_sheet(sheet_name)

    cols = primary_keys + [f"{col}{suffix}" for col in data_cols]
    cols = [c for c in cols if c in df.columns]

    # Headers
    for col_num, col_name in enumerate(cols, start=1):
        display_name = col_name.replace(suffix, "")
        ws.cell(1, col_num, display_name)
        _apply_header_style(ws.cell(1, col_num), fill_color=header_color, text_color=TextColor.WHITE)

    # Data rows
    _write_data_rows(ws, df, cols, start_row=2, max_rows=10000)

    _auto_adjust_columns(ws, len(cols))


def _write_reject_section(ws, df: pl.DataFrame, title: str, start_row: int) -> int:
    """Write a reject section to worksheet and return next available row."""
    # Title
    ws.cell(start_row, 1, title).font = Font(bold=True, size=12, color=TextColor.RED)
    row = start_row + 1

    # Headers
    _write_header_row(ws, list(df.columns), row_num=row, fill_color=HeaderColor.PEACH)
    row += 1

    # Data rows
    rows_written = _write_data_rows(ws, df, list(df.columns), start_row=row, max_rows=10000)

    return row + rows_written + 2


def _create_summary_sheet(
    wb: Workbook, summary_stats: dict, primary_keys: list[str], data_cols: list[str]
) -> None:
    """Create summary sheet with statistics and insights."""
    ws = wb.create_sheet("Summary", 0)

    # Title
    ws["A1"] = "Data File Validation Report"
    ws["A1"].font = Font(size=16, bold=True, color=TextColor.DARK_BLUE)
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(italic=True)

    # Summary Statistics
    row = 4
    ws[f"A{row}"] = "Summary Statistics"
    ws[f"A{row}"].font = Font(size=14, bold=True)
    row += 1

    # Header
    ws[f"A{row}"] = "Metric"
    ws[f"B{row}"] = "Value"
    for cell in [ws[f"A{row}"], ws[f"B{row}"]]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color=HeaderColor.LIGHT_BLUE, end_color=HeaderColor.LIGHT_BLUE, fill_type="solid"
        )
    row += 1

    # Add statistics
    stats_mapping = {
        "source_count": "Total Source Rows",
        "target_count": "Total Target Rows",
        "matching_rows": "Matching Rows",
        "mismatch_rows": "Mismatch Rows",
        "missing_in_source": "Missing in Source",
        "missing_in_target": "Missing in Target",
        "match_percentage": "Match Percentage",
        "data_match_percentage": "Data Match Percentage",
    }

    for key, label in stats_mapping.items():
        if key in summary_stats:
            ws[f"A{row}"] = label
            value = summary_stats[key]
            if isinstance(value, float) and key.endswith("_percentage"):
                ws[f"B{row}"] = f"{value:.2f}%"
            else:
                ws[f"B{row}"] = value
            row += 1

    row += 1

    # Data Quality Insights
    ws[f"A{row}"] = "Data Quality Insights"
    ws[f"A{row}"].font = Font(size=14, bold=True)
    row += 1

    ws[f"A{row}"] = "Primary Keys"
    ws[f"B{row}"] = ", ".join(primary_keys)
    row += 1

    ws[f"A{row}"] = "Columns Compared"
    ws[f"B{row}"] = len(data_cols)
    row += 1

    ws[f"A{row}"] = "Column Names"
    ws[f"B{row}"] = ", ".join(data_cols[:10]) + ("..." if len(data_cols) > 10 else "")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40


def _create_dual_comparison_sheet(
    wb: Workbook, df: pl.DataFrame, primary_keys: list[str], data_cols: list[str]
) -> None:
    """Create source-target dual comparison sheet."""
    ws = wb.create_sheet("Source-Target Comparison")

    source_cols = primary_keys + [f"{col}_source" for col in data_cols]
    target_cols = primary_keys + [f"{col}_target" for col in data_cols]

    source_cols = [c for c in source_cols if c in df.columns]
    target_cols = [c for c in target_cols if c in df.columns]

    col = 1
    ws.cell(1, col, "SOURCE DATA").font = Font(bold=True, size=12, color=TextColor.WHITE)
    ws.cell(1, col).fill = PatternFill(
        start_color=HeaderColor.DARK_BLUE, end_color=HeaderColor.DARK_BLUE, fill_type="solid"
    )

    for header in source_cols:
        col += 1
        display_name = header.replace("_source", "")
        ws.cell(2, col, display_name).font = Font(bold=True)
        ws.cell(2, col).fill = PatternFill(
            start_color=HeaderColor.LIGHT_BLUE, end_color=HeaderColor.LIGHT_BLUE, fill_type="solid"
        )

    col += 2  # Separator

    ws.cell(1, col, "TARGET DATA").font = Font(bold=True, size=12, color=TextColor.WHITE)
    ws.cell(1, col).fill = PatternFill(
        start_color=HeaderColor.GREEN, end_color=HeaderColor.GREEN, fill_type="solid"
    )

    for header in target_cols:
        col += 1
        display_name = header.replace("_target", "")
        ws.cell(2, col, display_name).font = Font(bold=True)
        ws.cell(2, col).fill = PatternFill(
            start_color=HeaderColor.LIGHT_GREEN, end_color=HeaderColor.LIGHT_GREEN, fill_type="solid"
        )

    # Data rows
    row = 3
    for record in df.iter_rows(named=True):
        col = 2
        for col_name in source_cols:
            ws.cell(row, col, record.get(col_name, ""))
            col += 1

        col += 2  # Separator

        for col_name in target_cols:
            ws.cell(row, col, record.get(col_name, ""))
            col += 1

        row += 1
        if row > 10003:  # Limit to 10000 data rows
            break

    for col_num in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15


def _create_source_data_sheet(
    wb: Workbook, df: pl.DataFrame, primary_keys: list[str], data_cols: list[str]
) -> None:
    """Create source data only sheet."""
    _create_comparison_sheet(
        wb, "Source Data", df, primary_keys, data_cols, suffix="_source", header_color=HeaderColor.DARK_BLUE
    )


def _create_target_data_sheet(
    wb: Workbook, df: pl.DataFrame, primary_keys: list[str], data_cols: list[str]
) -> None:
    """Create target data only sheet."""
    _create_comparison_sheet(
        wb, "Target Data", df, primary_keys, data_cols, suffix="_target", header_color=HeaderColor.GREEN
    )


def _create_mismatch_details_sheet(
    wb: Workbook, df: pl.DataFrame, primary_keys: list[str], data_cols: list[str]
) -> None:
    """Create sheet with only mismatched rows."""
    # Filter for mismatches
    if "validation_status" in df.columns:
        mismatch_df = df.filter(pl.col("validation_status").str.contains("Found in Both"))
    else:
        return

    if mismatch_df.is_empty():
        return

    ws = wb.create_sheet("Mismatch Details")

    # Title
    ws["A1"] = "Rows with Differences"
    ws["A1"].font = Font(size=14, bold=True, color=TextColor.RED)

    # Get all relevant columns
    all_cols = primary_keys + ["validation_status"]
    for col in data_cols:
        for suffix in ["_source", "_target", "_status"]:
            col_with_suffix = f"{col}{suffix}"
            if col_with_suffix in mismatch_df.columns:
                all_cols.append(col_with_suffix)

    all_cols = [c for c in all_cols if c in mismatch_df.columns]

    _write_header_row(ws, all_cols, row_num=3, fill_color=HeaderColor.PEACH)

    # Data rows
    row = 4
    for record in mismatch_df.iter_rows(named=True):
        for col_num, col_name in enumerate(all_cols, start=1):
            value = record.get(col_name, "")
            ws.cell(row, col_num, value)

            # Highlight mismatched cells
            if col_name.endswith("_status") and value == "MISMATCH":
                ws.cell(row, col_num).fill = PatternFill(
                    start_color=StatusColor.MISMATCH, end_color=StatusColor.MISMATCH, fill_type="solid"
                )

        row += 1
        if row > 10004:  # Limit to 10000 data rows
            break

    _auto_adjust_columns(ws, len(all_cols))


def _create_header_comparison_sheet(wb: Workbook, header_comparison: list[dict]) -> None:
    """Create header comparison sheet."""
    _create_line_comparison_sheet(wb, "Header Comparison", header_comparison)


def _create_trailer_comparison_sheet(wb: Workbook, trailer_comparison: list[dict]) -> None:
    """Create trailer comparison sheet."""
    _create_line_comparison_sheet(wb, "Trailer Comparison", trailer_comparison)


def _create_sample_data_sheet(wb: Workbook, sample_df: pl.DataFrame, primary_keys: list[str]) -> None:
    """Create sheet with sample matching data."""
    ws = wb.create_sheet("Sample Matching Data")

    ws["A1"] = "Sample of Matching Records (100% Match)"
    _apply_header_style(ws["A1"], fill_color=HeaderColor.LIGHT_GREEN_MATCH, text_color=TextColor.GREEN)
    ws["A1"].font = Font(size=12, bold=True, color=TextColor.GREEN)

    # Add a visible note of primary keys used for the comparison
    if primary_keys:
        ws["A2"] = "Primary Keys: " + ", ".join(primary_keys)
        ws["A2"].font = Font(italic=True, color=TextColor.DARK_BLUE)
    else:
        ws["A2"] = "Primary Keys: (none)"
        ws["A2"].font = Font(italic=True)

    # Shift header and data rows down to account for the primary keynote
    _write_header_row(ws, list(sample_df.columns), row_num=4, fill_color=HeaderColor.LIGHT_BLUE)

    _write_data_rows(ws, sample_df, list(sample_df.columns), start_row=5, max_rows=1000)

    _auto_adjust_columns(ws, len(sample_df.columns))


def _create_rejects_sheet(
    wb: Workbook, source_rejects: pl.DataFrame | None, target_rejects: pl.DataFrame | None
) -> None:
    """Create sheet with rejected records."""
    ws = wb.create_sheet("Rejected Records")
    row = 1

    if source_rejects is not None and not source_rejects.is_empty():
        row = _write_reject_section(ws, source_rejects, "Source File - Rejected Records", row)

    if target_rejects is not None and not target_rejects.is_empty():
        _write_reject_section(ws, target_rejects, "Target File - Rejected Records", row)

    _auto_adjust_columns(ws)


def export_to_excel(
    df: pl.DataFrame,
    output_file: str | Path,
    primary_keys: list[str],
    sample_df: pl.DataFrame | None = None,
    header_comparison: list[dict[str, str | int]] | None = None,
    trailer_comparison: list[dict[str, str | int]] | None = None,
    source_rejects: pl.DataFrame | None = None,
    target_rejects: pl.DataFrame | None = None,
    **summary_stats: int | float | str,
) -> None:
    """
    Export validation results to Excel with multiple sheets.

    Args:
        df: Polars DataFrame with comparison results
        output_file: Path for output Excel file (.xlsx)
        primary_keys: List of primary key column names
        sample_df: Optional sample data for 100% match scenarios
        header_comparison: Header line comparison results
        trailer_comparison: Trailer line comparison results
        source_rejects: Rejected rows from source file
        target_rejects: Rejected rows from target file
        **summary_stats: Additional summary statistics
    """
    output_file = Path(output_file)

    if output_file.suffix.lower() != ".xlsx":
        output_file = output_file.with_suffix(".xlsx")

    logger.info("=" * 60)
    logger.info("Exporting to Excel: %s", output_file.resolve().as_posix())
    logger.info("=" * 60)

    wb = Workbook()
    wb.remove(wb.active)

    data_cols = _extract_data_columns(df, sample_df, primary_keys)

    # Sheet 1: Summary
    _create_summary_sheet(wb, summary_stats, primary_keys, data_cols)

    # Sheet 2: Source-Target Comparison (Dual Table)
    if not df.is_empty():
        _create_dual_comparison_sheet(wb, df, primary_keys, data_cols)

    # Sheet 3: Source Data
    if not df.is_empty():
        _create_source_data_sheet(wb, df, primary_keys, data_cols)

    # Sheet 4: Target Data
    if not df.is_empty():
        _create_target_data_sheet(wb, df, primary_keys, data_cols)

    # Sheet 5: Mismatch Details
    if not df.is_empty():
        _create_mismatch_details_sheet(wb, df, primary_keys, data_cols)

    # Sheet 6: Header Comparison (if present)
    if header_comparison:
        _create_header_comparison_sheet(wb, header_comparison)

    # Sheet 7: Trailer Comparison (if present)
    if trailer_comparison:
        _create_trailer_comparison_sheet(wb, trailer_comparison)

    # Sheet 8: Sample Matching Data (if present)
    if sample_df is not None and not sample_df.is_empty():
        _create_sample_data_sheet(wb, sample_df, primary_keys)

    # Sheet 9: Rejected Records (if present)
    if (source_rejects is not None and not source_rejects.is_empty()) or (
        target_rejects is not None and not target_rejects.is_empty()
    ):
        _create_rejects_sheet(wb, source_rejects, target_rejects)

    # Save workbook
    wb.save(str(output_file))
    logger.info("Excel file exported successfully with %d sheets", len(wb.sheetnames))
    logger.info("Sheets: %s", ", ".join(wb.sheetnames))


def build_and_save_excel_file(
    df: pl.DataFrame,
    primary_keys: list[str],
    output_file: str | Path,
    sample_df: pl.DataFrame | None = None,
    header_comparison: list[dict[str, str | int]] | None = None,
    trailer_comparison: list[dict[str, str | int]] | None = None,
    source_rejects: pl.DataFrame | None = None,
    target_rejects: pl.DataFrame | None = None,
    **summary_stats: int | float | str,
) -> None:
    """
    Build and save Excel file with validation results.

    This is an alias for export_to_excel() maintained for backward compatibility.

    Args:
        df: Polars DataFrame with comparison results
        primary_keys: List of primary key column names
        output_file: Path for output Excel file (.xlsx)
        sample_df: Optional sample data for 100% match scenarios
        header_comparison: Header line comparison results
        trailer_comparison: Trailer line comparison results
        source_rejects: Rejected rows from source file
        target_rejects: Rejected rows from target file
        **summary_stats: Additional summary statistics
    """
    export_to_excel(
        df=df,
        output_file=output_file,
        primary_keys=primary_keys,
        sample_df=sample_df,
        header_comparison=header_comparison,
        trailer_comparison=trailer_comparison,
        source_rejects=source_rejects,
        target_rejects=target_rejects,
        **summary_stats,
    )
